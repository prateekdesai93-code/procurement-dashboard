#!/usr/bin/env python3
"""
Build a 2025 vs 2026 Excel comparison workbook.
3 sheets: Qty Comparison, Spend Comparison, YoY Summary.
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import argparse

MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

C_NAVY  = "1E293B"; C_BLUE  = "1D4ED8"; C_BLUE_L = "DBEAFE"
C_TEAL  = "0F766E"; C_TEAL_L = "CCFBF1"
C_ORG   = "C2410C"; C_ORG_L  = "FED7AA"
C_PUR   = "7C3AED"; C_PUR_L  = "EDE9FE"
C_GREEN = "15803D"; C_GREEN_L = "DCFCE7"
C_RED   = "B91C1C"; C_RED_L  = "FEE2E2"
C_WHITE = "FFFFFF"; C_LIGHT  = "F8FAFC"; C_MUTED = "64748B"; C_BORDER = "CBD5E1"

def fill(c): return PatternFill("solid", fgColor=c)
def fnt(bold=False, size=10, color=C_NAVY, italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)
def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def sc(ws, r, c, val, bold=False, size=10, color=C_NAVY, bg=None, ha="left", nf=None, italic=False):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = fnt(bold, size, color, italic)
    cell.alignment = aln(ha, "center")
    if bg: cell.fill = fill(bg)
    if nf: cell.number_format = nf
    return cell

def load(path):
    df = pd.read_excel(path)
    df.columns = [c.strip().lower() for c in df.columns]
    date_col = next((c for c in df.columns if 'date' in c), None)
    if date_col:
        df['order date'] = pd.to_datetime(df[date_col], dayfirst=True, errors='coerce')
    df = df.dropna(subset=['order date'])
    df['month'] = df['order date'].dt.month
    return df

def make_product_rows(df1, df2):
    all_prods = sorted(set(df1['product name'].dropna()) | set(df2['product name'].dropna()))
    rows = []
    for prod in all_prods:
        s1 = df1[df1['product name'] == prod]
        s2 = df2[df2['product name'] == prod]
        row = {'product': prod}
        for m in range(1, 13):
            row[f'2025_m{m}_qty']   = int(s1[s1['month']==m]['quantity'].sum())
            row[f'2025_m{m}_spend'] = round(float(s1[s1['month']==m]['total'].sum()), 2)
            row[f'2026_m{m}_qty']   = int(s2[s2['month']==m]['quantity'].sum())
            row[f'2026_m{m}_spend'] = round(float(s2[s2['month']==m]['total'].sum()), 2)
        row['t25q']  = int(s1['quantity'].sum())
        row['t25s']  = round(float(s1['total'].sum()), 2)
        row['t26q']  = int(s2['quantity'].sum())
        row['t26s']  = round(float(s2['total'].sum()), 2)
        rows.append(row)
    return pd.DataFrame(rows)

def write_comparison_sheet(wb, df, title, is_qty, accent25, accent26):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C5"
    ws.column_dimensions['A'].width = 0.8
    ws.column_dimensions['B'].width = 34
    for c in range(3, 33):
        ws.column_dimensions[get_column_letter(c)].width = 10
    ws.row_dimensions[1].height = 6

    for col in range(1, 33): ws.cell(2, col).fill = fill(C_NAVY)
    sc(ws, 2, 2, title, bold=True, size=13, color=C_WHITE, bg=C_NAVY)
    ws.merge_cells("B2:N2"); ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 6

    for col in range(3, 16): ws.cell(4, col).fill = fill(accent25)
    sc(ws, 4, 3, "◀  2025  ▶", bold=True, size=10, color=C_WHITE, bg=accent25, ha="center")
    ws.merge_cells("C4:N4")
    sc(ws, 4, 15, "2025 TOTAL", bold=True, size=9, color=C_WHITE, bg=accent25, ha="center")
    for col in range(16, 30): ws.cell(4, col).fill = fill(accent26)
    sc(ws, 4, 16, "◀  2026  ▶", bold=True, size=10, color=C_WHITE, bg=accent26, ha="center")
    ws.merge_cells("P4:AA4")
    sc(ws, 4, 29, "2026 TOTAL", bold=True, size=9, color=C_WHITE, bg=accent26, ha="center")
    sc(ws, 4, 30, "YoY Δ",  bold=True, size=9, color=C_WHITE, bg=C_NAVY, ha="center")
    sc(ws, 4, 31, "YoY %",  bold=True, size=9, color=C_WHITE, bg=C_NAVY, ha="center")
    ws.cell(4,30).fill = fill(C_NAVY); ws.cell(4,31).fill = fill(C_NAVY)
    ws.row_dimensions[4].height = 20

    sc(ws, 5, 2, "PRODUCT", bold=True, size=9, color=C_WHITE, bg="0F172A", ha="center")
    for mi, m in enumerate(MONTHS):
        ws.cell(5, 3+mi).value = m; ws.cell(5, 3+mi).font = fnt(True, 9, C_WHITE)
        ws.cell(5, 3+mi).fill = fill("1E3A8A"); ws.cell(5, 3+mi).alignment = aln("center")
        ws.cell(5,16+mi).value = m; ws.cell(5,16+mi).font = fnt(True, 9, C_WHITE)
        ws.cell(5,16+mi).fill = fill("134E4A"); ws.cell(5,16+mi).alignment = aln("center")
    sc(ws, 5, 15, "TOTAL", bold=True, size=9, color=C_WHITE, bg="1E3A8A", ha="center")
    sc(ws, 5, 29, "TOTAL", bold=True, size=9, color=C_WHITE, bg="134E4A", ha="center")
    sc(ws, 5, 30, "Δ",     bold=True, size=9, color=C_WHITE, bg="0F172A", ha="center")
    sc(ws, 5, 31, "% Δ",   bold=True, size=9, color=C_WHITE, bg="0F172A", ha="center")
    ws.row_dimensions[5].height = 16

    data_start = 6
    sorted_df = df.sort_values('t25q' if is_qty else 't25s', ascending=False)
    for ri, (_, row) in enumerate(sorted_df.iterrows()):
        r = data_start + ri
        bg = C_LIGHT if ri % 2 == 0 else C_WHITE
        sc(ws, r, 2, row['product'], size=9, bg=bg)
        q_key = 'qty' if is_qty else 'spend'
        nf = '#,##0' if is_qty else 'R#,##0'
        bg25 = C_BLUE_L if is_qty else C_ORG_L
        bg26 = C_TEAL_L if is_qty else C_PUR_L
        for mi in range(12):
            v25 = row[f'2025_m{mi+1}_{q_key}']
            v26 = row[f'2026_m{mi+1}_{q_key}']
            for val, col, cbg in [(v25, 3+mi, bg25), (v26, 16+mi, bg26)]:
                cell = ws.cell(r, col, value=val if val > 0 else None)
                cell.font = fnt(False, 9); cell.alignment = aln("center")
                cell.fill = fill(cbg if val > 0 else bg)
                if val > 0: cell.number_format = nf
        t25 = row['t25q' if is_qty else 't25s']
        t26 = row['t26q' if is_qty else 't26s']
        for val, col, cbg, clr in [(t25,15,bg25,"1D4ED8"),(t26,29,bg26,"0F766E")]:
            c = ws.cell(r, col, value=val if val>0 else None)
            c.font = fnt(True, 9, clr); c.alignment = aln("center")
            c.fill = fill(cbg if val>0 else bg)
            if val>0: c.number_format = nf
        if t25 > 0 and t26 > 0:
            delta = t26 - t25; pct = delta / t25
            cd = ws.cell(r, 30, value=delta)
            cd.font = fnt(True, 9, C_GREEN if delta>=0 else C_RED)
            cd.alignment = aln("center"); cd.fill = fill(C_GREEN_L if delta>=0 else C_RED_L)
            cd.number_format = '+#,##0;-#,##0;-' if is_qty else '+R#,##0;-R#,##0;-'
            cp = ws.cell(r, 31, value=pct)
            cp.font = fnt(True, 9, C_GREEN if delta>=0 else C_RED)
            cp.alignment = aln("center"); cp.fill = fill(C_GREEN_L if delta>=0 else C_RED_L)
            cp.number_format = '+0.0%;-0.0%;-'
        ws.row_dimensions[r].height = 14

    # Grand total
    rt = data_start + len(sorted_df)
    ws.row_dimensions[rt].height = 20
    for col in range(2, 32): ws.cell(rt, col).fill = fill(C_NAVY)
    sc(ws, rt, 2, "GRAND TOTAL", bold=True, size=10, color=C_WHITE, bg=C_NAVY)
    q_key = 'qty' if is_qty else 'spend'
    nf = '#,##0' if is_qty else 'R#,##0'
    for mi in range(12):
        for col, yr in [(3+mi,'2025'),(16+mi,'2026')]:
            v = int(df[f'{yr}_m{mi+1}_{q_key}'].sum()) if is_qty else round(float(df[f'{yr}_m{mi+1}_{q_key}'].sum()),2)
            c = ws.cell(rt, col, value=v); c.font = fnt(True,9,C_NAVY)
            c.alignment = aln("center"); c.fill = fill("93C5FD" if yr=='2025' else "5EEAD4"); c.number_format = nf
    for col, key, clr in [(15,'t25q'if is_qty else't25s',"93C5FD"),(29,'t26q'if is_qty else't26s',"5EEAD4")]:
        v = int(df[key].sum()) if is_qty else round(float(df[key].sum()),2)
        c = ws.cell(rt, col, value=v); c.font = fnt(True,10,C_NAVY)
        c.alignment = aln("center"); c.fill = fill(clr); c.number_format = nf
    d = (int(df['t26q'].sum())-int(df['t25q'].sum())) if is_qty else round(float(df['t26s'].sum()-df['t25s'].sum()),2)
    c = ws.cell(rt,30,value=d); c.font = fnt(True,10,C_WHITE); c.alignment = aln("center"); c.fill = fill(C_NAVY)
    c.number_format = '+#,##0;-#,##0;-' if is_qty else '+R#,##0;-R#,##0;-'
    t25t = df['t25q'].sum() if is_qty else df['t25s'].sum()
    p = d/t25t if t25t else 0
    cp = ws.cell(rt,31,value=p); cp.font = fnt(True,10,C_WHITE); cp.alignment = aln("center")
    cp.fill = fill(C_NAVY); cp.number_format = '+0.0%;-0.0%;-'

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--file2025', required=True)
    parser.add_argument('--file2026', required=True)
    parser.add_argument('--output',   default='Comparison_2025_2026.xlsx')
    args = parser.parse_args()

    print("Loading files...")
    df1 = load(args.file2025)
    df2 = load(args.file2026)
    df  = make_product_rows(df1, df2)
    print(f"  Products: {len(df)}")

    wb = Workbook()
    wb.remove(wb.active)

    write_comparison_sheet(wb, df, "📦 Qty Comparison",   True,  "1D4ED8", "0F766E")
    write_comparison_sheet(wb, df, "💰 Spend Comparison", False, "9A3412", "4C1D95")

    # YoY Summary sheet
    ws3 = wb.create_sheet("📊 YoY Summary")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "B5"
    ws3.column_dimensions['A'].width = 0.8; ws3.column_dimensions['B'].width = 34
    for c,w in {3:14,4:14,5:12,6:12,7:14,8:14,9:12,10:12}.items():
        ws3.column_dimensions[get_column_letter(c)].width = w
    ws3.row_dimensions[1].height = 6
    for col in range(1,11): ws3.cell(2,col).fill = fill(C_NAVY)
    sc(ws3,2,2,"YEAR-ON-YEAR SUMMARY — PRODUCTS IN BOTH 2025 & 2026",bold=True,size=12,color=C_WHITE,bg=C_NAVY)
    ws3.merge_cells("B2:J2"); ws3.row_dimensions[2].height = 30
    ws3.row_dimensions[3].height = 6
    for ci,h in enumerate(["PRODUCT","2025 Qty","2026 Qty","Qty Δ","Qty %","2025 Spend","2026 Spend","Spend Δ","Spend %"]):
        sc(ws3,4,2+ci,h,bold=True,size=9,color=C_WHITE,bg="0F172A",ha="center")
    ws3.row_dimensions[4].height = 18
    both = df[(df['t25q']>0)&(df['t26q']>0)].sort_values('t25s',ascending=False)
    for ri,(_,row) in enumerate(both.iterrows()):
        r=5+ri; bg=C_LIGHT if ri%2==0 else C_WHITE
        sc(ws3,r,2,row['product'],size=9,bg=bg)
        for col,val,nf,clr,cbg in [
            (3,int(row['t25q']),'#,##0',"1D4ED8",C_BLUE_L),(4,int(row['t26q']),'#,##0',"0F766E",C_TEAL_L),
            (7,round(float(row['t25s']),2),'R#,##0',"9A3412",C_ORG_L),(8,round(float(row['t26s']),2),'R#,##0',"4C1D95",C_PUR_L),
        ]:
            c=ws3.cell(r,col,value=val); c.font=fnt(True,9,clr); c.alignment=aln("center"); c.fill=fill(cbg); c.number_format=nf
        dq=int(row['t26q']-row['t25q']); pq=dq/row['t25q'] if row['t25q'] else 0
        ds=round(float(row['t26s']-row['t25s']),2); ps=ds/row['t25s'] if row['t25s'] else 0
        for col,val,nf in [(5,dq,'+#,##0;-#,##0;-'),(6,pq,'+0.0%;-0.0%;-'),(9,ds,'+R#,##0;-R#,##0;-'),(10,ps,'+0.0%;-0.0%;-')]:
            c=ws3.cell(r,col,value=val); clr2=C_GREEN if val>=0 else C_RED
            c.font=fnt(True,9,clr2); c.alignment=aln("center"); c.fill=fill(C_GREEN_L if val>=0 else C_RED_L); c.number_format=nf
        ws3.row_dimensions[r].height = 14

    for ws in wb.worksheets:
        ws.sheet_properties.tabColor = "1D4ED8"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1

    wb.save(args.output)
    print(f"Saved: {args.output}")

if __name__ == '__main__':
    main()
